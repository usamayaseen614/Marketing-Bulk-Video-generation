"""
app.py — Streamlit UI for the bulk marketing video generator.

Run with:  streamlit run app.py
"""

import io
import logging
import shutil
import tempfile
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from streamlit import config as st_config

from preview_editor import preview_editor
from video_generator import (
    ALL_COLUMNS,
    CANVAS_H,
    CANVAS_W,
    CTA_VIDEO_SLOTS,
    FONT_CHOICES,
    FONT_CUSTOM,
    REQUIRED_COLUMNS,
    TEXT_STYLES,
    RenderConfig,
    RowResult,
    VideoGenerator,
    missing_optional_columns,
    validate_dataframe,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("app")

# Outputs land outside the per-run TemporaryDirectory so the download button
# keeps working after temp cleanup. Each browser session gets its own subfolder
# so concurrent users can't clobber each other's batches.
OUTPUT_ROOT = Path(tempfile.gettempdir()) / "bulk_video_generator"

# When static serving is enabled (production/Docker), oversized ZIPs are
# published here and streamed from disk by Tornado instead of being buffered
# in Python memory by st.download_button.
STATIC_ROOT = Path(__file__).parent / "static"
STATIC_DOWNLOADS = STATIC_ROOT / "downloads"

# Above this size the ZIP is not loaded into memory for the download button.
MAX_DOWNLOAD_BYTES = 5 * 1024 * 1024 * 1024

# Output folders from sessions idle longer than this are deleted on the next
# batch run, so a long-lived server never fills its disk.
STALE_RUN_HOURS = 24


# --------------------------------------------------------------------------- helpers

def get_session_id() -> str:
    """Stable id for this browser session, used to isolate output folders."""
    if "session_id" not in st.session_state:
        st.session_state["session_id"] = uuid.uuid4().hex[:12]
    return st.session_state["session_id"]


def cleanup_stale_dirs(root: Path, max_age_hours: int = STALE_RUN_HOURS) -> None:
    """Delete other sessions' output folders once they go stale."""
    if not root.is_dir():
        return
    cutoff = time.time() - max_age_hours * 3600
    for child in root.iterdir():
        try:
            if child.is_dir() and child.stat().st_mtime < cutoff:
                shutil.rmtree(child, ignore_errors=True)
        except OSError:
            pass


@dataclass
class Workspace:
    """Uploaded assets materialized on disk inside a temp directory."""
    bg_dir: Path
    video_path: Path
    cta_path: Path
    font_path: Optional[Path]
    work_dir: Path
    cta_video_slots: list = field(default_factory=list)


def build_workspace(tmp: Path, video_file, zip_file, cta_file, font_file,
                    cta_video_slot_files=None) -> Workspace:
    """Write the in-memory uploads to disk where FFmpeg/PIL can read them."""
    video_path = tmp / "input.mp4"
    video_path.write_bytes(video_file.getvalue())

    cta_path = tmp / "cta.png"
    cta_path.write_bytes(cta_file.getvalue())

    bg_dir = tmp / "backgrounds"
    with zipfile.ZipFile(io.BytesIO(zip_file.getvalue())) as zf:
        zf.extractall(bg_dir)

    font_path = None
    if font_file is not None:
        font_path = tmp / f"custom_font{Path(font_file.name).suffix.lower()}"
        font_path.write_bytes(font_file.getvalue())

    # One sub-folder per CTA slot, keeping each sample's original filename so the
    # Excel CTA_Clip_<n> cells can pin one by name.
    cta_video_slots = []
    for i, files in enumerate(cta_video_slot_files or [], start=1):
        slot_paths = []
        if files:
            slot_dir = tmp / f"cta_slot_{i}"
            slot_dir.mkdir(parents=True, exist_ok=True)
            for f in files:
                p = slot_dir / Path(f.name).name
                p.write_bytes(f.getvalue())
                slot_paths.append(p)
        cta_video_slots.append(slot_paths)

    work_dir = tmp / "work"
    work_dir.mkdir(exist_ok=True)
    return Workspace(bg_dir, video_path, cta_path, font_path, work_dir, cta_video_slots)


def make_generator(ws: Workspace, config: RenderConfig, output_dir: Path) -> VideoGenerator:
    config.font_path = str(ws.font_path) if ws.font_path else None
    return VideoGenerator(
        config=config,
        bg_dir=ws.bg_dir,
        video_path=ws.video_path,
        cta_path=ws.cta_path,
        work_dir=ws.work_dir,
        output_dir=output_dir,
        cta_video_slots=ws.cta_video_slots,
    )


def apply_saved_edits(df: pd.DataFrame, edits: dict[int, dict]) -> pd.DataFrame:
    """Overlay values saved from the preview editor onto the uploaded sheet.
    Keys are 1-based data row numbers (1 = first row below the header)."""
    out = df.copy()
    for row_no, cols in edits.items():
        if not 1 <= int(row_no) <= len(out):
            continue
        for col, value in cols.items():
            # Empty optional columns parse as float64; pandas refuses to put
            # a string (e.g. a color hex) into them — widen to object first.
            if (
                isinstance(value, str)
                and col in out.columns
                and out[col].dtype != object
            ):
                out[col] = out[col].astype(object)
            out.loc[out.index[int(row_no) - 1], col] = value
    return out


def updated_excel_bytes(excel_bytes: bytes, edits: dict[int, dict]) -> bytes:
    """The original workbook with the saved values written into the first
    sheet — everything else (formatting, formulas elsewhere, extra sheets)
    is preserved. Columns the sheet doesn't have yet are appended after the
    last used column."""
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.worksheets[0]
    headers = {
        str(cell.value).strip(): cell.column for cell in ws[1] if cell.value is not None
    }
    next_col = (max(headers.values()) + 1) if headers else 1
    for row_no, cols in sorted(edits.items()):
        for name, value in cols.items():
            col = headers.get(name)
            if col is None:
                col = next_col
                ws.cell(row=1, column=col, value=name)
                headers[name] = col
                next_col += 1
            ws.cell(row=int(row_no) + 1, column=col, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_render_log(path: Path, results: list[RowResult]) -> None:
    ok = sum(r.ok for r in results)
    lines = [
        f"Bulk video render log — {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"Total rows: {len(results)} | Succeeded: {ok} | Failed: {len(results) - ok}",
        "-" * 70,
    ]
    for r in sorted(results, key=lambda r: r.row_number):
        status = "OK    " if r.ok else "FAILED"
        lines.append(f"Row {r.row_number:4d}  {status}  {r.filename or ''}")
        for w in r.warnings:
            lines.append(f"           warning: {w}")
        if r.error:
            lines.append(f"           error: {r.error}")
    path.write_text("\n".join(lines), encoding="utf-8")


def package_zip(run_dir: Path, results: list[RowResult]) -> Path:
    """Bundle all successful MP4s + the render log. ZIP_STORED because MP4s
    are already compressed — recompressing wastes minutes for ~0% gain."""
    zip_path = run_dir / "marketing_videos.zip"
    log_path = run_dir / "render_log.txt"
    write_render_log(log_path, results)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_STORED) as zf:
        for r in results:
            if r.ok and r.output_path and r.output_path.is_file():
                zf.write(r.output_path, r.filename)
        zf.write(log_path, "render_log.txt")
    return zip_path


def run_batch(df: pd.DataFrame, generator: VideoGenerator, workers: int) -> list[RowResult]:
    """Render every row with a live progress bar. FFmpeg does the heavy lifting
    in subprocesses, so a thread pool is enough for parallelism."""
    progress = st.progress(0.0, text="Starting renders…")
    results: list[RowResult] = []

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [
            pool.submit(generator.render_row, idx + 1, row)
            for idx, (_, row) in enumerate(df.iterrows())
        ]
        for future in as_completed(futures):
            results.append(future.result())
            done = len(results)
            ok = sum(r.ok for r in results)
            progress.progress(
                done / len(futures),
                text=f"Rendered {done}/{len(futures)} — {ok} ok, {done - ok} failed",
            )
    progress.empty()
    return sorted(results, key=lambda r: r.row_number)


def show_results(results: list[RowResult]) -> None:
    ok = [r for r in results if r.ok]
    failed = [r for r in results if not r.ok]

    if not failed:
        st.success(f"All {len(ok)} videos rendered successfully.")
    elif ok:
        st.warning(f"{len(ok)} videos rendered, {len(failed)} failed.")
    else:
        st.error(f"All {len(failed)} rows failed.")

    if failed:
        with st.expander(f"Failed rows ({len(failed)})", expanded=True):
            st.dataframe(
                pd.DataFrame(
                    [{"Row": r.row_number, "Error": r.error} for r in failed]
                ),
                hide_index=True,
                width="stretch",
            )

    warnings = [(r.row_number, w) for r in results for w in r.warnings]
    if warnings:
        with st.expander(f"Warnings ({len(warnings)})"):
            for row_number, message in warnings:
                st.text(f"Row {row_number}: {message}")


def offer_download(zip_path: Path) -> None:
    size = zip_path.stat().st_size
    if size <= MAX_DOWNLOAD_BYTES:
        st.caption(f"ZIP size: {size / 1024 / 1024:.1f} MB — saved at `{zip_path}`")
        st.download_button(
            "⬇️ Download all videos (ZIP)",
            data=zip_path.open("rb"),
            file_name="marketing_videos.zip",
            mime="application/zip",
            type="primary",
        )
        return

    # st.download_button buffers the whole file in memory per click — too risky
    # for multi-GB batches. With static serving on (the production/Docker
    # setup), move the ZIP under ./static and let Tornado stream it from disk.
    if st_config.get_option("server.enableStaticServing"):
        dest = STATIC_DOWNLOADS / get_session_id() / zip_path.name
        if zip_path != dest:
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(zip_path), dest)
            st.session_state["result_zip"] = str(dest)
        st.caption(f"ZIP size: {size / 1024 / 1024:.1f} MB")
        st.markdown(
            f"### [⬇️ Download all videos (ZIP)](/app/static/downloads/{get_session_id()}/{dest.name})"
        )
    else:
        st.info(
            "The ZIP is too large to stream through the browser reliably. "
            f"Grab it directly from disk:\n\n`{zip_path}`"
        )


# --------------------------------------------------------------------------- UI

if st_config.get_option("server.enableStaticServing"):
    STATIC_DOWNLOADS.mkdir(parents=True, exist_ok=True)

st.set_page_config(page_title="Bulk Video Generator", page_icon="🎬", layout="wide")
st.title("🎬 Bulk Marketing Video Generator")
st.caption(
    "Upload an Excel sheet, a promo video, background images, and a CTA image — "
    "get one 1080x1920 (9:16) MP4 per row, ready for Reels, TikTok, and Shorts."
)

# ---- sidebar: layout & output configuration
with st.sidebar:
    st.header("⚙️ Settings")

    st.subheader("Video placement")
    st.caption(
        "Defaults for every row — the Excel columns `Video_X`, `Video_Y`, "
        "`Video_Width`, `Video_Height` override them per video."
    )
    randomize_video = st.checkbox(
        "Randomize position per video",
        value=False,
        help="Each video gets its own random spot (avoiding the CTA and any "
             "explicitly positioned texts). Reproducible per row, so the "
             "preview matches the final render.",
    )
    if randomize_video:
        video_x, video_y = 0, 0  # ignored; per-row positions are computed
    else:
        video_x = st.number_input("Video X", 0, CANVAS_W, 90)
        video_y = st.number_input("Video Y", 0, CANVAS_H, 300)
    video_w = st.number_input("Video width", 50, CANVAS_W, 900)
    video_h = st.number_input("Video height", 50, CANVAS_H, 900)

    st.subheader("CTA image placement")
    st.caption(
        "Defaults for every row — the Excel columns `CTA_X`, `CTA_Y`, "
        "`CTA_Width`, `CTA_Height` override them per video."
    )
    cta_x = st.number_input("CTA X", 0, CANVAS_W, 340)
    cta_y = st.number_input("CTA Y", 0, CANVAS_H, 1600)
    cta_w = st.number_input("CTA width", 10, CANVAS_W, 400)
    cta_h = st.number_input("CTA height", 10, CANVAS_H, 160)
    cta_fade_start = st.number_input(
        "CTA fade-in start (s)", 0.0, 30.0, 1.0, 0.1,
        help="The CTA image is invisible until this time, then fades in. "
             "Per-row override: `CTA_Fade_Start`.",
    )
    cta_fade_duration = st.number_input(
        "CTA fade-in duration (s)", 0.0, 30.0, 0.5, 0.1,
        help="How long the fade-in takes. Per-row override: `CTA_Fade_Duration`.",
    )

    st.subheader("CTA video (optional)")
    st.caption(
        f"A sequence of {CTA_VIDEO_SLOTS} clips that always play in order "
        "(1 → 2 → 3 → 4) in one shared box. Each clip is a *pool* of sample "
        "videos (up to ~30): one is chosen per output video — pinned by an Excel "
        "`CTA_Clip_<n>` cell, otherwise picked at random. Leave all empty to skip."
    )
    cta_video_slot_files = [
        st.file_uploader(
            f"Clip {i} — sample videos (MP4)", type=["mp4"],
            accept_multiple_files=True, key=f"cta_clip_{i}",
            help="One of these is chosen per output video for this position.",
        )
        for i in range(1, CTA_VIDEO_SLOTS + 1)
    ]
    cta_video_x = st.number_input("CTA video X", 0, CANVAS_W, 360)
    cta_video_y = st.number_input("CTA video Y", 0, CANVAS_H, 1200)
    cta_video_w = st.number_input("CTA video width", 50, CANVAS_W, 360)
    cta_video_h = st.number_input("CTA video height", 50, CANVAS_H, 360)
    cta_video_fade_start = st.number_input(
        "CTA video fade-in start (s)", 0.0, 30.0, 0.5, 0.1,
        help="Per-row override: `CTA_Video_Fade_Start`.",
    )
    cta_video_fade_duration = st.number_input(
        "CTA video fade-in duration (s)", 0.0, 30.0, 0.5, 0.1,
        help="Per-row override: `CTA_Video_Fade_Duration`.",
    )
    st.caption(
        "Playback speed per clip position — 1 = normal, 2 = twice as fast, "
        "0.5 = half. Per-row overrides: `CTA_Video_Speed_<n>` for one clip, or "
        "`CTA_Video_Speed` for the whole row."
    )
    cta_video_speeds = [
        st.number_input(
            f"Clip {i} speed (×)", 0.25, 4.0, 1.0, 0.05, key=f"cta_speed_{i}",
        )
        for i in range(1, CTA_VIDEO_SLOTS + 1)
    ]

    st.subheader("Text style")
    st.caption(
        "Defaults for every text — per-row Excel columns `*_Font` and `*_Style` "
        "override them, and `*_BgColor` adds a highlight box behind any text."
    )
    default_font = st.selectbox(
        "Default font", FONT_CHOICES, index=0,
        help="A bundled font family, the system font, or your uploaded font. "
             "Override per text with `Headline_Font` etc.",
    )
    default_style = st.selectbox(
        "Default artistic style", TEXT_STYLES, index=0,
        help="classic = plain · outline = contrasting border · shadow = drop "
             "shadow · neon = glow. Override per text with `Headline_Style` etc.",
    )

    st.subheader("Output")
    crf = st.slider("Quality (CRF — lower = better/bigger)", 16, 28, 18)
    preset = st.select_slider(
        "Encoder speed",
        options=["veryfast", "faster", "fast", "medium", "slow"],
        value="medium",
        help="Faster presets render quicker but produce slightly larger files.",
    )
    workers = st.slider(
        "Parallel renders", 1, 4, 2,
        help="Concurrent FFmpeg processes. 2 is a good default on office machines.",
    )
    font_file = st.file_uploader(
        "Custom font (TTF/OTF, optional)", type=["ttf", "otf"],
        help=f"Upload your own font, then pick “{FONT_CUSTOM}” as the default "
             "font above (or set a `*_Font` cell to it) to use it.",
    )

config = RenderConfig(
    video_x=int(video_x), video_y=int(video_y),
    video_w=int(video_w), video_h=int(video_h),
    cta_x=int(cta_x), cta_y=int(cta_y),
    cta_w=int(cta_w), cta_h=int(cta_h),
    cta_fade_start=float(cta_fade_start), cta_fade_duration=float(cta_fade_duration),
    cta_video_x=int(cta_video_x), cta_video_y=int(cta_video_y),
    cta_video_w=int(cta_video_w), cta_video_h=int(cta_video_h),
    cta_video_fade_start=float(cta_video_fade_start),
    cta_video_fade_duration=float(cta_video_fade_duration),
    cta_video_speeds=[float(s) for s in cta_video_speeds],
    default_font=default_font, default_style=default_style,
    crf=int(crf), preset=preset,
    randomize_video_pos=randomize_video,
)

# ---- uploads
st.subheader("1. Upload assets")
col1, col2 = st.columns(2)
with col1:
    excel_file = st.file_uploader("Excel file (.xlsx)", type=["xlsx"])
    video_file = st.file_uploader("Promo video (MP4) — used in every output", type=["mp4"])
with col2:
    zip_file = st.file_uploader("Background images (ZIP)", type=["zip"])
    cta_file = st.file_uploader("CTA image (PNG)", type=["png"])

# ---- Excel validation & preview table
df = None
if excel_file is not None:
    try:
        df = pd.read_excel(io.BytesIO(excel_file.getvalue()), engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not read the Excel file: {exc}")

    if df is not None:
        missing = validate_dataframe(df)
        if missing:
            st.error(
                "The Excel file is missing required columns: "
                + ", ".join(f"`{c}`" for c in missing)
            )
            with st.expander("Supported columns"):
                st.code(
                    "Required:\n  " + "\n  ".join(REQUIRED_COLUMNS)
                    + "\n\nOptional (defaults / auto-placement when absent):\n  "
                    + "\n  ".join(c for c in ALL_COLUMNS if c not in REQUIRED_COLUMNS)
                )
            df = None
        elif df.empty:
            st.error("The Excel file has no data rows.")
            df = None
        else:
            st.success(f"Excel loaded — {len(df)} video(s) to generate.")
            # Edits saved from the preview editor belong to one specific
            # sheet — drop them (and the stale preview) on a new upload.
            file_key = f"{excel_file.name}:{excel_file.size}"
            if st.session_state.get("excel_file_key") != file_key:
                st.session_state["excel_file_key"] = file_key
                for stale in ("row_edits", "preview_payload", "preview_nonce",
                              "preview_row", "preview_baseline_edits"):
                    st.session_state.pop(stale, None)
            row_edits = st.session_state.get("row_edits") or {}
            if row_edits:
                df = apply_saved_edits(df, row_edits)
            absent = missing_optional_columns(df)
            if absent:
                st.caption(
                    "Columns not in this sheet (defaults / auto-placement will be used): "
                    + ", ".join(f"`{c}`" for c in absent)
                )
            with st.expander("Preview spreadsheet data"):
                st.dataframe(df, hide_index=True, width="stretch")

ready = df is not None and video_file is not None and zip_file is not None and cta_file is not None
if not ready:
    st.info("Upload all four files to enable preview and generation.")

# ---- actions
st.subheader("2. Generate")
col_row, col_preview, col_generate = st.columns([1, 2, 2], vertical_alignment="bottom")
preview_row = col_row.number_input(
    "Row to preview", 1, len(df) if df is not None else 1, 1, disabled=not ready,
    help="Excel data row number (1 = the first row below the header).",
)
preview_clicked = col_preview.button("👁️ Preview Row", disabled=not ready, width="stretch")
generate_clicked = col_generate.button(
    "🚀 Generate All Videos", disabled=not ready, type="primary", width="stretch"
)

if preview_clicked and ready:
    with st.spinner(f"Rendering preview of row {preview_row}…"):
        with tempfile.TemporaryDirectory(prefix="bvg_preview_") as tmp:
            try:
                ws = build_workspace(Path(tmp), video_file, zip_file, cta_file,
                                     font_file, cta_video_slot_files)
                generator = make_generator(ws, config, Path(tmp) / "out")
                # Same deterministic background assignment as the real batch,
                # so the preview shows the row's actual background.
                df_preview, _ = generator.assign_backgrounds(df)
                payload = generator.build_editor_payload(df_preview.iloc[int(preview_row) - 1])
                st.session_state["preview_payload"] = payload
                st.session_state["preview_nonce"] = uuid.uuid4().hex
                st.session_state["preview_row"] = int(preview_row)
                # The payload was built WITH this row's saved edits applied,
                # so a later save reports changes relative to them — keep the
                # snapshot to merge against (and to detect reverts).
                st.session_state["preview_baseline_edits"] = dict(
                    (st.session_state.get("row_edits") or {}).get(int(preview_row), {})
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("Preview failed")
                st.session_state.pop("preview_payload", None)
                st.error(f"Preview failed: {exc}")

if "preview_payload" in st.session_state and not generate_clicked:
    st.caption(
        f"Row {st.session_state.get('preview_row', 1)} preview (1080x1920) — drag the "
        "video, CTA, or texts to move them, drag the corner handle to resize, recolor "
        "texts, and add a background box; click “Save to Excel” in the panel to apply the "
        "changed values to that row for previews, generation, and the Excel download below."
    )
    saved = preview_editor(
        st.session_state["preview_payload"], st.session_state["preview_nonce"]
    )
    # The component echoes its last value on every rerun — only apply a save
    # that belongs to the current preview (nonce) and hasn't been seen (token).
    if (
        isinstance(saved, dict)
        and saved.get("nonce") == st.session_state.get("preview_nonce")
        and saved.get("token") != st.session_state.get("editor_save_token")
    ):
        st.session_state["editor_save_token"] = saved.get("token")
        merged = dict(st.session_state.get("preview_baseline_edits", {}))
        merged.update(saved.get("values") or {})
        edits = st.session_state.setdefault("row_edits", {})
        row_no = int(st.session_state["preview_row"])
        if merged:
            edits[row_no] = merged
        else:
            edits.pop(row_no, None)
        # Rerun so everything above (data preview, generation input) reflects
        # the just-saved values in this same interaction.
        st.rerun()

row_edits = st.session_state.get("row_edits") or {}
if row_edits and excel_file is not None and not generate_clicked:
    rows_txt = ", ".join(str(r) for r in sorted(row_edits))
    st.success(
        f"Saved edits for row(s) {rows_txt} — applied to previews and generation. "
        "Download the updated sheet to keep your Excel in sync."
    )
    col_dl, col_clear = st.columns([2, 1], vertical_alignment="center")
    col_dl.download_button(
        "⬇️ Download updated Excel",
        data=updated_excel_bytes(excel_file.getvalue(), row_edits),
        file_name=excel_file.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if col_clear.button("🗑️ Discard saved edits"):
        st.session_state.pop("row_edits", None)
        st.rerun()

if generate_clicked and ready:
    st.session_state.pop("result_zip", None)
    session_id = get_session_id()
    # Reap stale output from idle sessions, then wipe only THIS session's
    # previous run — concurrent users keep their batches intact.
    cleanup_stale_dirs(OUTPUT_ROOT)
    cleanup_stale_dirs(STATIC_DOWNLOADS)
    session_root = OUTPUT_ROOT / session_id
    shutil.rmtree(session_root, ignore_errors=True)
    shutil.rmtree(STATIC_DOWNLOADS / session_id, ignore_errors=True)
    run_dir = session_root / time.strftime("run_%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)

    started = time.time()
    # The TemporaryDirectory holds uploads + per-row PNGs and is deleted
    # automatically when the batch finishes (success or failure).
    with tempfile.TemporaryDirectory(prefix="bvg_run_") as tmp:
        try:
            ws = build_workspace(Path(tmp), video_file, zip_file, cta_file,
                                 font_file, cta_video_slot_files)
            generator = make_generator(ws, config, run_dir / "videos")
            df_run, bg_warnings = generator.assign_backgrounds(df)
            for message in bg_warnings:
                st.warning(message)
            results = run_batch(df_run, generator, workers)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Batch failed")
            st.error(f"Batch failed before rendering could start: {exc}")
            results = []

    if results:
        show_results(results)
        elapsed = time.time() - started
        st.caption(f"Finished in {elapsed:.0f}s ({elapsed / len(results):.1f}s per video).")
        if any(r.ok for r in results):
            st.session_state["result_zip"] = str(package_zip(run_dir, results))

if "result_zip" in st.session_state:
    zip_path = Path(st.session_state["result_zip"])
    if zip_path.is_file():
        offer_download(zip_path)
